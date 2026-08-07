#!/usr/bin/env python3
"""
parse_theirstack_sheet.py — robust, tab-targeted reader for TheirStack pipeline sheets.

WHY THIS EXISTS
  The Jobs and Decision-Makers tabs share ONE spreadsheet. Two facts broke the
  old daily run:
    1. Drive `read_file_content` renders Decision-Makers FIRST, so parsing "the
       first table" read the stale DM tab, not Jobs.
    2. `read_file_content` caps its output (~107K chars), so it never reaches the
       newest Jobs rows on a large sheet. (SDR: it showed 165 of 797 real rows.)
  Result: the run saw "0 new jobs" while 600+ jobs sat unenriched, so NO decision
  makers were found.

THE FIX — read the COMPLETE workbook, address the tab by NAME:
  Call Drive `download_file_content` with
    exportMimeType = application/vnd.openxmlformats-officedocument.spreadsheetml.sheet
  The result overflows, so the harness saves it to a file as
    {"content": "<base64 xlsx>", "id", "mimeType", "title"}
  Pass that saved file here. openpyxl reads the named worksheet in full — no
  truncation, no tab-order ambiguity, no base64 round-trip through the agent.
  (Fallback: if given a read_file_content dump {"fileContent": "<markdown>"} or
  raw markdown, it parses the pipe-tables instead — but that path is truncation-
  prone and only used if no xlsx download is available.)

USAGE
  python3 parse_theirstack_sheet.py <saved_file> <Tab> <state.json> [--key KEY] [--out FILE]
    <Tab>  "Jobs" (key=Job Posting, state=processed_jobs)
           "Decision-Makers" (key=Email, state=processed_dms)
    --out  Write the full new_rows array as JSON here (keeps stdout small).

STDOUT  Compact summary (counts, date span, unique new companies, sample rows).
"""
import json, base64, io, re, sys, os, argparse


def canon(s):
    return re.sub(r'[^a-z0-9]', '', str(s).lower())


# ---------- input loading: xlsx-download (preferred) or markdown dump ----------

def _rows_from_xlsx_bytes(xb, tab):
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(xb), read_only=True, data_only=True)
    wsname = next((n for n in wb.sheetnames if canon(n) == canon(tab)), None)
    if not wsname:
        return None, wb.sheetnames
    ws = wb[wsname]
    rows = [[("" if c is None else str(c)).strip() for c in r]
            for r in ws.iter_rows(values_only=True)]
    rows = [r for r in rows if any(c for c in r)]
    return (rows[0], rows[1:]) if rows else ([], [])


def load_rows(path, tab):
    """Return (header, data_rows) for the requested tab. Accepts ANY of:
       - the saved download JSON {"content": "<base64 xlsx>"} (preferred)
       - a raw .xlsx file (e.g. if the download was decoded first)
       - a read_file_content dump {"fileContent": "<markdown>"} or raw markdown.
    """
    raw = open(path, "rb").read()
    # Raw xlsx (zip magic) — guards the trap of feeding a decoded workbook.
    if raw[:4] == b"PK\x03\x04":
        return _rows_from_xlsx_bytes(raw, tab)
    obj = None
    try:
        obj = json.loads(raw.decode("utf-8", "replace"))
    except Exception:
        obj = None
    if isinstance(obj, dict) and obj.get("content"):
        return _rows_from_xlsx_bytes(base64.b64decode(obj["content"]), tab)
    content = obj["fileContent"] if isinstance(obj, dict) and "fileContent" in obj \
        else raw.decode("utf-8", "replace")
    return _markdown_table(content, tab)


def _markdown_table(content, tab):
    def cells(line):
        line = line.strip().strip("|")
        out = []
        for c in re.split(r'(?<!\\)\|', line):
            c = c.strip()
            for a, b in (("\\|", "|"), ("\\_", "_"), ("\\&", "&"), ("\\-", "-")):
                c = c.replace(a, b)
            out.append(c)
        return out

    def is_sep(l):
        return re.fullmatch(r'[\s|:\-]+', l.strip()) is not None and '-' in l

    tables, cur = [], []
    for ln in content.splitlines():
        if ln.lstrip().startswith("|"):
            cur.append(ln)
        elif cur:
            tables.append(cur); cur = []
    if cur:
        tables.append(cur)
    want = canon(tab)
    for t in tables:
        rs = [cells(l) for l in t if not is_sep(l)]
        rs = [r for r in rs if any(c.strip() for c in r)]
        if len(rs) < 2:
            continue
        hdr = [canon(h) for h in rs[0]]
        if (want.startswith("job") and "datepulled" in hdr and "jobposting" in hdr) or \
           (not want.startswith("job") and "firstname" in hdr and "email" in hdr):
            return rs[0], rs[1:]
    return None, [r[0][:6] for r in tables if r]


# ---------------------------------- main ----------------------------------

def norm_url(u):
    u = (u or "").strip().replace("\\_", "_").replace("\\&", "&").replace("\\-", "-")
    return u.split("?")[0].rstrip("/").lower()


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("read_file")
    ap.add_argument("tab")
    ap.add_argument("state_file")
    ap.add_argument("--key", default=None)
    ap.add_argument("--out", default=None)
    a = ap.parse_args()

    header, data = load_rows(a.read_file, a.tab)
    if header is None:
        print(json.dumps({"error": "tab not found", "tab": a.tab, "tabs_seen": data}))
        sys.exit(2)
    header = [h.strip() for h in header]

    is_jobs = canon(a.tab).startswith("job")
    key = a.key or ("Job Posting" if is_jobs else "Email")
    kidx = next((i for i, h in enumerate(header) if canon(h) == canon(key)), None)
    if kidx is None:
        print(json.dumps({"error": "key column not found", "key": key, "header": header}))
        sys.exit(2)
    didx = next((i for i, h in enumerate(header)
                 if canon(h) in ("datepulled", "dateposted", "dateadded")), None)
    cidx = next((i for i, h in enumerate(header) if canon(h) in ("website", "domain", "company")), None)

    state = json.load(open(a.state_file)) if os.path.exists(a.state_file) else {}
    slist = "processed_jobs" if is_jobs else "processed_dms"
    keyfn = norm_url if is_jobs else (lambda x: str(x).strip().lower())
    processed = {keyfn(x) for x in state.get(slist, [])}

    new, seen, already, dates, companies = [], set(), 0, [], set()
    for r in data:
        if len(r) <= kidx or not r[kidx].strip():
            continue
        if didx is not None and len(r) > didx and r[didx].strip():
            dates.append(r[didx][:10])
        nk = keyfn(r[kidx])
        if nk in processed:
            already += 1
            continue
        if nk in seen:
            continue
        seen.add(nk)
        if cidx is not None and len(r) > cidx and r[cidx].strip():
            companies.add(r[cidx].strip().lower())
        new.append({(header[i] if i < len(header) else f"col{i}"): (r[i] if i < len(r) else "")
                    for i in range(max(len(header), len(r)))})

    if a.out:
        json.dump(new, open(a.out, "w"), ensure_ascii=False)

    keep = {canon(header[0]), canon("Company"), canon("Job Title"), canon("Country"), canon(header[kidx])}
    samples = [{k: v for k, v in row.items() if canon(k) in keep} for row in new[:8]]
    print(json.dumps({
        "tab": a.tab, "key_column": header[kidx], "state_list": slist,
        "total_data_rows": len(data), "already_in_state": already, "new_count": len(new),
        "unique_new_companies": len(companies),
        "tab_oldest_date": min(dates) if dates else None,
        "tab_newest_date": max(dates) if dates else None,
        "out_file": a.out, "sample_new_rows": samples,
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
